
# 与excel表交互，遍历excel 中的接口测试用例

import  sys,os
sys.path.append('/Users/low/PycharmProjects/snowFramework/src')
sys.path.append('/Users/low/PycharmProjects/snowFramework/commons')

import  dresponse
from  dresponse import DreponseF
import openpyxl,json,re
import requests
import dlog
from dlog import SnowLog

class operateExcel:
    def operatExcel(self,excelUrl,sheetName,checkPoint):
        wb = openpyxl.load_workbook(excelUrl)
        ws = wb[sheetName]
        for i in range(2, ws.max_row + 1):
            if ws.cell(row=i, column=10).value == 'Yes':
                gurl = ws.cell(row=i, column=3).value
                gpath = ws.cell(row=i, column=4).value
                gmethod = ws.cell(row=i, column=5).value
                gtype = ws.cell(row=i, column=6).value
                gparams = ws.cell(row=i, column=7).value
                gcheckPoint = ws.cell(row=i, column=8).value

                fullPath = gurl + gpath
                gmethod = gmethod.lower()

                if gmethod == 'post' and gtype == 'Form':
                    headers = {"content-type": "application/x-www-form-urlencoded"}
                elif gmethod == 'post' and gtype == 'Json':
                    headers = {"content-type": "application/json"}
                else:
                    headers = None

                gparams = eval(gparams)

                snow = DreponseF()

                r = snow.Dreponse(gmethod, fullPath, headers, gparams,gcheckPoint)

                r_status = r[0]
                r_json = r[1]
                r_totaltime = r[2]

                r_jsonS = json.dumps(r_json,ensure_ascii=False)

                snowLog = SnowLog()
                snowLog.snowFileLog('INFO', os.getcwd() + '/logFile/log.log')

                snowLog.snowInfo(r_status)

                snowLog.snowInfo(r_json)
                snowLog.snowInfo(type(r_json))
                snowLog.snowInfo(r_jsonS)
                snowLog.snowInfo(type(r_jsonS))
                snowLog.snowInfo(r_totaltime)


                error_code = r_json[checkPoint]
                snowD = {}
                snowD[checkPoint] = error_code
                # print(snowD, type(snowD))
                snowDs = json.dumps(snowD)

                # print('实际结果：', snowDs, type(snowDs))
                # print('预期结果', gcheckPoint, type(gcheckPoint))

                # a = print('实际结果：', snowDs, type(snowDs))
                # yuqi =


                snowLog.snowInfo(snowDs)
                snowLog.snowInfo(type(snowDs))
                snowLog.snowInfo(gcheckPoint)
                snowLog.snowInfo(type(gcheckPoint))


                if r_status == 200:
                    if re.search(gcheckPoint, snowDs):

                        snowLog.snowInfo('测试结果通过')
                        testR = '测试结果通过'
                    else:

                        snowLog.snowInfo('测试结果不通过')
                        testR = '测试结果不通过'

                else:
                    # print('响应出错啦，没有返回200')
                    snowLog.snowError('响应出错啦，没有返回200')


                ws.cell(row = i, column = 11,value = testR)
                ws.cell(row = i, column = 12,value = r_jsonS)
                ws.cell(row = i, column = 13,value = r_totaltime)
                wb.save('/Users/low/PycharmProjects/snowFramework/reports/testcase_student.xlsx')


            else:
                continue





# get响应有中文，响应数据体dict 类型：json {'error_code': 0, 'reason': 'successed', 'books': ['小强测试品牌', '小强python自动化测试', '小强软件测试疯狂讲义']} <class 'dict'>
# 使用json.dumps(r_json)，响应数据体str类型： {"error_code": 0, "reason": "successed", "books": ["\u5c0f\u5f3a\u6d4b\u8bd5\u54c1\u724c", "\u5c0f\u5f3apython\u81ea\u52a8\u5316\u6d4b\u8bd5", "\u5c0f\u5f3a\u8f6f\u4ef6\u6d4b\u8bd5\u75af\u72c2\u8bb2\u4e49"]} <class 'str'>
# 问题：转成字符串，其中中文，无法正常显示。 google
# 解决：json.dumps(r_json,ensure_ascii=False) 中增加ensure_ascii=False，中文正常显示，写入excel表格，中文正常显示
# 响应数据体str类型： {"error_code": 0, "reason": "successed", "books": ["小强测试品牌", "小强python自动化测试", "小强软件测试疯狂讲义"]} <class 'str'>

# 疑惑：excel表格取出 预期结果 "error_code": 0 <class 'str'>
# 响应返回数据体中拿到json 数据，放置新字典中，将字典用json.dunps转成字符串 实际结果： {"error_code": 0} <class 'str'>
# re.search 对比成功了 ？

# wb = openpyxl.load_workbook('/Users/liguizeng/Desktop/testcase_student.xlsx')
# ws = wb['测试详情']
#
# print( ws.cell(row = 2,column =10).value)
# for i in range(2, ws.max_row +1):
#     if ws.cell(row = i ,column = 10).value == 'Yes':
#         gurl = ws.cell(row = i ,column = 3).value
#         gpath = ws.cell(row = i ,column = 4).value
#         gmethod = ws.cell(row = i ,column = 5).value
#         gtype =  ws.cell(row = i ,column = 6).value
#         gparams = ws.cell(row = i ,column = 7).value
#         gcheckPoint = ws.cell(row = i ,column = 8).value
#
#         fullPath = gurl + gpath
#         gmethod = gmethod.lower()
#         print('%%%%',gmethod)
#         if gmethod == 'post' and gtype == 'Form':
#             headers = {"content-type": "application/x-www-form-urlencoded"}
#         elif gmethod == 'post' and gtype == 'Json':
#             headers = {"content-type": "application/json"}
#         else:
#             headers= None
# #
#         print('哎哈哈',gmethod,fullPath,headers,gparams)
#         print(type(gparams),type(headers))
#         # gparams = json.loads(gparams)
#         gparams =eval(gparams)
#         print(gparams,type(gparams))
#
#         r = snow.snowRequests(gmethod, fullPath, headers, gparams)
#
#将下面死的接口，从excel 表中拿出来请求。 请求通过，再放置class中，将能提炼出来的参数变成参数传递。
#  最后去main文件中，将该py文件地址引入，类创建，参数在main中定义



#  接口地址是死的，先请求通过
# url = 'http://localhost:10000'
# pathG = '/book_list'
# pathGP = '/book_info'
# pathPostF = '/login1'
# pathPostJ = '/login2'
# fullPathG = url +pathG
# fullPathGP = url +pathGP
# fullPathPF = url + pathPostF
# fullPathPJ = url + pathPostJ
# #
# # # get 无参数
# method = 'get'
# r = snow.snowRequests(method,fullPathG,None,None)

# get有参数
# queries = {"bookname":"小强软件测试疯狂讲义","checkstatus":"on"}
# r = snow.snowRequests(method,fullPathGP,None,queries)

# # post，form
# headers = {"content-type":"application/x-www-form-urlencoded"}
# params = {"username":"xiaoqiang","pwd":"123123"}
# method = 'post'
# r = snow.snowRequests(method,fullPathPF,headers = headers,params=params)

#
# post Json
# headers = {"content-type":"application/json"}
# json = {"username":"xiaoqiang","pwd":"123123"}
# r = snow.snowRequests(method,fullPathPJ,headers=headers,params = json)
#
